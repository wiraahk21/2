# Import modul-modul yang diperlukan
import customtkinter as ctk
import openpyxl as xl
from tkinter import *
from tkinter import messagebox
import pandas as pd
from sklearn.ensemble import RandomForestClassifier

# Membaca data dari file CSV menggunakan pandas
data = pd.read_csv('MUSTAHIK3.csv')

# Mengubah pendapatan menjadi label penerima zakat (1 jika pendapatan < 600 ribu, 0 jika tidak)
data['penerima_zakat'] = (data['pendapatan'] < 600000).astype(int)

# Memisahkan fitur (X) dan label (y)
X = data[['pendapatan']]
y = data['penerima_zakat']

# Membangun model Random Forest
rf_model = RandomForestClassifier(n_estimators=100, random_state=42)
rf_model.fit(X, y)

# Membuat DataFrame yang berisi penerima zakat dengan pendapatan terendah
penerima_zakat_terendah = data[data['penerima_zakat'] == 1]
rekomendasi = penerima_zakat_terendah[['nama', 'pendapatan', 'alamat']]

# Mengurutkan DataFrame berdasarkan pendapatan (dari terkecil ke terbesar)
rekomendasi = rekomendasi.sort_values(by='pendapatan')

# Membuat instance dari kelas CTk (custom tkinter)
app = ctk.CTk()
app.geometry("800x500")
app.title("Identifikasi Penerima Zakat Fitrah dengan Pemanfaatan AI")

# Menentukan font yang akan digunakan dalam program
font1 = ("Arial", 20, "bold")
font2 = ("Arial", 14, "bold")

# Fungsi untuk menampilkan halaman berikutnya
def show_next_page(current_page, next_page):
    current_page.pack_forget()
    next_page.pack()

# Fungsi untuk menampilkan halaman sebelumnya
def show_previous_page(current_page, previous_page):
    current_page.pack_forget()
    previous_page.pack()

# Fungsi untuk menyimpan data ke file Excel "MUZZAKI.xlsx"
def submit():
    file=xl.load_workbook("MUZZAKI.xlsx")
    sheet=file["Sheet1"]
    Nama_value=Nama_entry.get()
    Usia_value=Usia_entry.get()
    Kelamin_value=Kelamin_combobox.get()
    TLP_value=TLP_entry.get()
    Alamat_value=Alamat_entry.get()
    sheet.cell(column=1, row=sheet.max_row+1, value=Nama_value)
    sheet.cell(column=2, row=sheet.max_row, value=Usia_value)
    sheet.cell(column=3, row=sheet.max_row, value=Kelamin_value)
    sheet.cell(column=4, row=sheet.max_row, value=TLP_value)
    sheet.cell(column=5, row=sheet.max_row, value=Alamat_value)
    file.save("MUZZAKI.xlsx")
    messagebox.showinfo(title="sukses", message="Data telah direkam")

# Fungsi untuk menghapus data yang dimasukkan di widget-entry
def clear():
    Nama_entry.delete(0,END)
    Usia_entry.delete(0,END)
    TLP_entry.delete(0,END)
    Alamat_entry.delete(0,END)

# Fungsi untuk menghasilkan rekomendasi penerima zakat
def generate_rekomendasi():
    global rekomendasi
    if not rekomendasi.empty:
        data_prediksi = rekomendasi[['pendapatan']]

        # Prediksi menggunakan model Random Forest
        prediksi = rf_model.predict(data_prediksi)
        if prediksi[0] == 1:
            nama_penerima = rekomendasi.iloc[0]['nama']
            alamat_penerima = rekomendasi.iloc[0]['alamat']
            label_rekomendasi.configure(text=f"Rekomendasi Penerima Zakat:\n Nama: {nama_penerima} \n Alamat: {alamat_penerima}", text_color="white", font=("Arial", 18))
        else:
            label_rekomendasi.configure(text="Tidak ada rekomendasi penerima zakat dengan pendapatan terendah.", text_color="white", font=("Arial", 18))
    else:
        label_rekomendasi.configure(text="Tidak ada rekomendasi penerima zakat dengan pendapatan terendah.", text_color="white", font=("Arial", 18))

# Fungsi untuk menangani konfirmasi "Ya"
def konfirmasi_ya():
    if not rekomendasi.empty:
        nama_penerima = rekomendasi.iloc[0]['nama']
        alamat_penerima = rekomendasi.iloc[0]['alamat']
        messagebox.showinfo("Konfirmasi", f"Penerima zakat anda adalah sebagai berikut. \nNama: {nama_penerima} \nAlamat: {alamat_penerima}\nSilahkan untuk dicatat datanya.")
        rekomendasi.drop(rekomendasi.index[0], inplace=True)
        generate_rekomendasi()

# Fungsi untuk menangani konfirmasi "Tidak"
def konfirmasi_tidak():
    if not rekomendasi.empty:
        rekomendasi.drop(rekomendasi.index[0], inplace=True)
        generate_rekomendasi()
    else:
        messagebox.showinfo("Konfirmasi", "Tidak ada rekomendasi penerima zakat lagi.")

# Halaman 1: Memasukkan data muzzaki
page1 = ctk.CTkFrame(master=app, width=800, height=500)
page1.pack()

# Membuar label halaman awal
Awal_label = ctk.CTkLabel(page1, text="Selamat datang di aplikasi sistem rekomendasi\n penerima zakat. Silahkan masukkan data anda\n untuk pencatatan data muzzaki.", text_color="white", font=("Arial", 20), width=10)
Awal_label.place(x=190, y=60)

# Membuat label Nama:
Nama_label = ctk.CTkLabel(page1, text="Nama:", text_color="white", font=font1, width=10)
Nama_label.place(x=170, y=150)

# Membuat label Usia:
Usia_label = ctk.CTkLabel(page1, text="Usia:", text_color="White", font=font1, width=10)
Usia_label.place(x=170, y=200)

# Membuat label Kelamin:
Kelamin_label = ctk.CTkLabel(page1, text="Jenis Kelamin:", text_color="White", font=font1, width=10)
Kelamin_label.place(x=170, y=250)

# Membuat label Telepon:
TLP_label = ctk.CTkLabel(page1, text="No. Tlpn:", text_color="white", font=font1, width=10)
TLP_label.place(x=170, y=300)

# Membuat label Alamat:
Alamat_label = ctk.CTkLabel(page1, text="Alamat:", text_color="white", font=font1, width=10)
Alamat_label.place(x=170, y=350)

# Membuat entry nama
Nama_entry_var = ctk.StringVar()
Nama_entry = ctk.CTkEntry(page1, font=font2, textvariable=Nama_entry_var, width=300)
Nama_entry.place(x=315, y=150)

# Membuat entry usia
Usia_entry_var = ctk.StringVar()
Usia_entry = ctk.CTkEntry(page1, font=font2, textvariable=Usia_entry_var, width=300)
Usia_entry.place(x=315, y=200)

# Membuat menu pilihan jenis kelamin
Kelamin_combobox = ctk.CTkComboBox(page1, font=font2, dropdown_hover_color="#145e15", values=["Laki-laki", "Perempuan"], width=300)
Kelamin_combobox.place(x=315, y=250)

# Membuat entry telepon
TLP_entry_var = ctk.StringVar()
TLP_entry = ctk.CTkEntry(page1, font=font2, textvariable=TLP_entry_var, width=300)
TLP_entry.place(x=315, y=300)

# Membuat entry alamat
Alamat_entry_var = ctk.StringVar()
Alamat_entry = ctk.CTkEntry(page1, font=font2, textvariable=Alamat_entry_var, width=300)
Alamat_entry.place(x=315, y=350)

# Membuat tombol untuk memasukan data ke file MUZZAKI.xlsx
submit_button=ctk.CTkButton(page1, command=submit, font=font1, text="Submit", fg_color="#0a8018", hover_color="#0a8018")
submit_button.place(x=200,y=400)

# Membuat tombol untuk menghapus widget
clear_button=ctk.CTkButton(page1, command=clear, font=font1, text="Clear", fg_color="#ad6507", hover_color="#ad6507")
clear_button.place(x=450,y=400)

# Membuat tombol untuk ke halaman 2
next_button1 = ctk.CTkButton(page1, text="Selanjutnya", font=font1, command=lambda: show_next_page(page1, page2))
next_button1.place(x=480,y=450)

# Halaman 2: Menampilkan rekomendasi penerima zakat
page2 = ctk.CTkFrame(master=app, width=800, height=500)
page2.pack()

# Membuat label halaman awal di halaman 2
Nama1_label = ctk.CTkLabel(page2, text="Halo Muzzaki, ini adalah menu dimana \n anda mendapatkan rekomendasi penerima zakat. \n Klik tombol Generate dibawah ini.", text_color="white", font=("Arial", 18), width=10)
Nama1_label.place(x=200, y=50)

# Membuat tombol generate untuk mendapatkan informasi penerima zakat
generate_button = ctk.CTkButton(page2, text="Generate", font=font1, command=generate_rekomendasi)
generate_button.place(x=330, y=150)

# Membuat label rekomendasi data penerima zakat dari sistem
label_rekomendasi = ctk.CTkLabel(page2, text="")
label_rekomendasi.place(x=60, y=200)

# Membuat tombol Ya untuk mengonfirmasi apakah user memilih orang tersebut
button_ya = ctk.CTkButton(page2, text="Ya", font=font1, fg_color="#0a8018", hover_color="#0a8018", command=konfirmasi_ya)
button_ya.place(x=200, y=300)

# Membuat tombol Tidak untuk melakukan regenerate lagi
button_tidak = ctk.CTkButton(page2, text="Tidak", font=font1,fg_color="#ad6507", hover_color="#ad6507", command=konfirmasi_tidak)
button_tidak.place(x=460, y=300)

# Membuat tombol selanjutnya untuk pindah ke halaman 3
next_button2 = ctk.CTkButton(page2, font=font1, text="Selanjutnya", command=lambda: show_next_page(page2, page3))
next_button2.place(x=480,y=450)

# Membuat tombol kembali untuk pindah ke halaman 1
previous_button2 = ctk.CTkButton(page2, font=font1, text="Kembali", command=lambda: show_previous_page(page2, page1))
previous_button2.place(x=180,y=450)

# Halaman 3: Menampilkan niat zakat fitrah
page3 = ctk.CTkFrame(master=app, width=800, height=500)
page3.pack()

# Membuat label halaman awal pada halaman 3
Nama_label = ctk.CTkLabel(page3, text="Sebelum membayar zakat fitrah anda, yuk baca niat \n zakat fitrah dahulu.", text_color="white", font=("Arial", 18), width=10)
Nama_label.place(x=190, y=50)

# Membuat label niat zakat fitrah
Nama2_label = ctk.CTkLabel(page3, text="ﻧَﻮَﻳْﺖُ أَﻥْ أُﺧْﺮِﺝَ ﺯَﻛَﺎﺓَ ﺍﻟْﻔِﻄْﺮِ ﻋَﻦْ ﻧَﻔْسيْ ﻓَﺮْﺿًﺎ ِﻟﻠﻪِ ﺗَﻌَﺎﻟَﻰ", text_color="white", font=("Arial", 23), width=10)
Nama2_label.place(x=200, y=120)

# Membuat label arti dari niat zakat fitrah
Nama3_label = ctk.CTkLabel(page3, text="Artinya: “Aku niat mengeluarkan zakat fitrah untuk \n diriku sendiri, fardu karena Allah Ta‘âlâ.”", text_color="white", font=("Arial", 18), width=10)
Nama3_label.place(x=200, y=180)

# Membuat label untuk identitas pembuat sistem ini
Nama4_label = ctk.CTkLabel(page3, text="Powered by", text_color="white", font=("Arial", 12), width=10)
Nama4_label.place(x=365, y=380)
Nama5_label = ctk.CTkLabel(page3, text="PTOIR 2021", text_color="white", font=("Arial", 30, "bold"), width=10)
Nama5_label.place(x=315, y=400)

# Membuat tombol kembali untuk pindah ke halaman 2
previous_button3 = ctk.CTkButton(page3, font=font1, text="Kembali", command=lambda: show_previous_page(page3, page2))
previous_button3.place(x=180,y=450)

# Membuat tombol selesai untuk keluar dari sistem
finish_button = ctk.CTkButton(page3, font=font1, text="Selesai", command=app.quit)
finish_button.place(x=480,y=450)

# Memulai aplikasi tkinter
app.mainloop()